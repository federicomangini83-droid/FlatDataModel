import pandas as pd
import openpyxl
import datetime

# Registra l'ora di inizio
start_time = datetime.datetime.now()
print("Ora di inizio:", start_time)

class DatiColonne:
    def __init__(self, IdTabCol=None,ValCol=None,NumCol=None):
        self.IdTabCol = IdTabCol
        self.ValCol = ValCol
        self.NumCol = NumCol
        
class DatiRighe:
    def __init__(self, IdTabRow=None,ValRow=None,NumRow=None):
        self.IdTabRow = IdTabRow
        self.ValRow = ValRow
        self.NumRow = NumRow

class MaxCol:
    def __init__(self, IdTabMaxC=None,MaxC=None):
        self.IdTabMaxC = IdTabMaxC
        self.MaxC = MaxC
        
class MaxRow:
    def __init__(self, IdTabMaxR=None,MaxR=None):
        self.IdTabMaxR = IdTabMaxR
        self.MaxR = MaxR
        
class MappingCell:
    def __init__(self, IdTabMapping=None,RigaMapping=None,ColonnaMapping=None,Cod_Conto=None,Cod_Dest2=None,Cod_Dest3=None,Cod_Dest4=None,Cod_Dest5=None,Cod_Categoria=None,Formula=None,Sign=None,EBASign=None,CC=None,Coordinate=None):
        self.IdTabMapping = IdTabMapping
        self.RigaMapping = RigaMapping
        self.ColonnaMapping = ColonnaMapping
        self.Cod_Conto = Cod_Conto
        self.Cod_Dest2 = Cod_Dest2
        self.Cod_Dest3 = Cod_Dest3
        self.Cod_Dest4 = Cod_Dest4
        self.Cod_Dest5 = Cod_Dest5
        self.Cod_Categoria = Cod_Categoria
        self.Formula = Formula
        self.Sign = Sign
        self.EBASign = EBASign
        self.CC = CC
        self.Coordinate = Coordinate

DatiColonneList = []
DatiRigheList = []
MaxColList = []
MaxRowList = []
MappingList = []

valore_riga = None
valore_colonna = None
Puntatore_instance = None

####################################################################################

# PARAMETRI DA MODIFICARE

workbook_path=r"C:\Users\federico.mangini\Downloads\DOWNLOAD\DataModelFlat_Finrep.xlsx"
save_path=r"C:\Users\federico.mangini\Downloads\DOWNLOAD"
start_name=r'FR_'
end_name=r'_420'

####################################################################################

filename=r'TGK_MAPPING'

workbook = openpyxl.load_workbook(workbook_path)
lista_sheets = workbook.sheetnames
sheet_stile_riferimento = lista_sheets[0]
sheet_stile_riferimento = workbook[sheet_stile_riferimento]
cella_stile_riferimento = sheet_stile_riferimento['A1']
cella_stile_val = sheet_stile_riferimento['A2']
cella_stile_val_col = sheet_stile_riferimento['B1']
stile_riferimento = cella_stile_riferimento._style
stile_rif_val = cella_stile_val._style
stile_rif_val_col = cella_stile_val_col._style
stile_rif_val2 = cella_stile_riferimento.fill.start_color.rgb


for nome_sheet in lista_sheets:
    countC=1
    countR=1
    sheet = workbook[nome_sheet]
    for rigas in sheet.iter_rows():
        for cella in rigas:
            colonna=cella.column
            if colonna==1:
                continue
            elif cella.value is not None and (cella._style==stile_rif_val_col or cella.fill.start_color.rgb==stile_rif_val2):
                countC=countC+1
                DatiColInstance = DatiColonne(
                    IdTabCol=nome_sheet,
                    ValCol=str(cella.value).zfill(4),
                    NumCol=countC
                )
                DatiColonneList.append(DatiColInstance)
            elif cella.value is None:
                break
        break
    MaxColInstance = MaxCol(
            IdTabMaxC=nome_sheet,
            MaxC=countC
    )
    MaxColList.append(MaxColInstance)
    for columns in sheet.iter_cols():
        for cella in columns:
            riga=cella.row
            if riga==1:
                continue
            elif cella.value is not None and (cella._style==stile_rif_val or cella.fill.start_color.rgb==stile_rif_val2):
                countR=countR+1
                DatiRowInstance = DatiRighe(
                    IdTabRow=nome_sheet,
                    ValRow=str(cella.value).zfill(4),
                    NumRow=countR
                )
                DatiRigheList.append(DatiRowInstance)
            elif cella.value is None:
                break
        break
    MaxRowInstance = MaxRow(
            IdTabMaxR=nome_sheet,
            MaxR=countR
    )
    MaxRowList.append(MaxRowInstance)

DatiColonneTab = {
    'IdTabCol': [DatiColInstance.IdTabCol for DatiColInstance in DatiColonneList],
    'ValCol': [DatiColInstance.ValCol for DatiColInstance in DatiColonneList],
    'NumCol': [DatiColInstance.NumCol for DatiColInstance in DatiColonneList]
}
    
DatiRigheTab = {
    'IdTabRow': [DatiRowInstance.IdTabRow for DatiRowInstance in DatiRigheList],
    'ValRow': [DatiRowInstance.ValRow for DatiRowInstance in DatiRigheList],
    'NumRow': [DatiRowInstance.NumRow for DatiRowInstance in DatiRigheList]
}
    
MaxColTab = {
    'IdTabMaxC': [MaxColInstance.IdTabMaxC for MaxColInstance in MaxColList],
    'MaxC': [MaxColInstance.MaxC for MaxColInstance in MaxColList]
}
    
MaxRowTab = {
    'IdTabMaxR': [MaxRowInstance.IdTabMaxR for MaxRowInstance in MaxRowList],
    'MaxR': [MaxRowInstance.MaxR for MaxRowInstance in MaxRowList]
}




df = pd.DataFrame(DatiColonneTab)
#nome_file_csv1=r'C:\Users\federico.mangini\Downloads\provaColonne1.csv'
#df.to_csv(nome_file_csv1, index=False)   
DatiColonneTab = df.to_dict(orient='records')

df = pd.DataFrame(DatiRigheTab)
#nome_file_csv2=r'C:\Users\federico.mangini\Downloads\provaRighe1.csv'
#df.to_csv(nome_file_csv2, index=False) 
DatiRigheTab = df.to_dict(orient='records')

df = pd.DataFrame(MaxColTab)
#nome_file_csv3=r'C:\Users\federico.mangini\Downloads\provaMaxColonne1.csv'
#df.to_csv(nome_file_csv3, index=False)  
MaxColTab = df.to_dict(orient='records') 

df = pd.DataFrame(MaxRowTab)
#nome_file_csv4=r'C:\Users\federico.mangini\Downloads\provaMaxRighe1.csv'
#df.to_csv(nome_file_csv4, index=False)  
MaxRowTab = df.to_dict(orient='records') 

for nome_sheet in lista_sheets:
    sheet = workbook[nome_sheet]
    for record in MaxColTab:
        if record.get('IdTabMaxC')==nome_sheet:
            indCol=record.get('MaxC')
            break
    for record in MaxRowTab:
        if record.get('IdTabMaxR')==nome_sheet:
            indRow=record.get('MaxR')
            break  
    for rigas in sheet.iter_rows(min_row=2, max_row=indRow, min_col=2, max_col=indCol):
        for cella in rigas:
            if cella.fill.fgColor.rgb!=stile_rif_val2:
                mapping=cella.value
                if mapping is not None:
                    mapping=str(mapping)
                    if mapping.find('Account= ''') != -1 and mapping.find('Formula= ''') == -1 and mapping.find('''Dest 2 = ''') != -1 and mapping.find('Calculation logic= ''') == -1:
                        mapping=mapping.replace('\n','')
                        lenMapping=len(mapping)
                        posConto = mapping.find('Account= ''')
                        posDest2 = mapping.find('''Dest 2 = ''')
                        posDest3 = mapping.find('''Dest 3 = ''')
                        posDest4 = mapping.find('''Dest 4 = ''')
                        posDest5 = mapping.find('''Dest 5 = ''')
                        posCategoria = mapping.find('''Category= ''')
                        posSign = mapping.find('''DB Storage Sign= ''')
                        posEBASign = mapping.find('''EBA Sign= ''')
                        lenConto = len('Account= ''')
                        lenDest2 = len('''Dest 2 = ''')
                        lenDest3 = len('''Dest 3 = ''')
                        lenDest4 = len('''Dest 4 = ''')
                        lenDest5 = len('''Dest 5 = ''')
                        lenCategoria = len('''Category= ''')
                        lenSign = len('''DB Storage Sign= ''')
                        lenEBASign = len('''EBA Sign= ''')
                        Conto=mapping[posConto+lenConto+1:posDest2-1].strip()
                        Dest2=mapping[posDest2+lenDest2+1:posDest3-1].strip()
                        Dest3=mapping[posDest3+lenDest3+1:posDest4-1].strip()
                        Dest4=mapping[posDest4+lenDest4+1:posDest5-1].strip()
                        Dest5=mapping[posDest5+lenDest5+1:posCategoria-1].strip()
                        Categoria=mapping[posCategoria+lenCategoria+1:posSign-1].strip()
                        Formula1=None
                        CC1=None
                        Sign1=mapping[posSign+lenSign+1:posEBASign-1].strip()
                        EBASign1=mapping[posEBASign+lenEBASign+1:lenMapping-1].strip()
                    elif mapping.find('Account= ''') == -1 and mapping.find('Formula= ''') != -1 and mapping.find('Calculation logic= ''') == -1:
                        mapping=mapping.replace('\n','')
                        lenMapping=len(mapping)
                        posFormula = mapping.find('Formula= ''')
                        posSign = mapping.find('''DB Storage Sign= ''')
                        posEBASign = mapping.find('''EBA Sign= ''')
                        lenFormula = len('Formula= ''')
                        lenSign = len('''DB Storage Sign= ''')
                        lenEBASign = len('''EBA Sign= ''')
                        Conto=None
                        Dest2=None
                        Dest3=None
                        Dest4=None
                        Dest5=None
                        Categoria=None
                        Formula1= mapping[posFormula+lenFormula+1:posSign-1].strip()
                        CC1=None
                        Sign1=mapping[posSign+lenSign+1:posEBASign-1].strip()
                        EBASign1=mapping[posEBASign+lenEBASign+1:lenMapping-1].strip()
                    elif mapping.find('Account= ''') != -1 and mapping.find('Formula= ''') == -1 and mapping.find('''Dest 2 = ''') != -1 and mapping.find('Calculation logic= ''') != -1:
                        mapping=mapping.replace('\n','')
                        lenMapping=len(mapping)
                        posConto = mapping.find('Account= ''')
                        posDest2 = mapping.find('''Dest 2 = ''')
                        posDest3 = mapping.find('''Dest 3 = ''')
                        posDest4 = mapping.find('''Dest 4 = ''')
                        posDest5 = mapping.find('''Dest 5 = ''')
                        posCategoria = mapping.find('''Category= ''')
                        posSign = mapping.find('''DB Storage Sign= ''')
                        posEBASign = mapping.find('''EBA Sign= ''')
                        posCC= mapping.find('''Calculation logic= ''')
                        lenConto = len('Account= ''')
                        lenDest2 = len('''Dest 2 = ''')
                        lenDest3 = len('''Dest 3 = ''')
                        lenDest4 = len('''Dest 4 = ''')
                        lenDest5 = len('''Dest 5 = ''')
                        lenCategoria = len('''Category= ''')
                        lenSign = len('''DB Storage Sign= ''')
                        lenEBASign = len('''EBA Sign= ''')
                        lenCC = len('''Calculation logic= ''')
                        Conto=mapping[posConto+lenConto+1:posDest2-1].strip()
                        Dest2=mapping[posDest2+lenDest2+1:posDest3-1].strip()
                        Dest3=mapping[posDest3+lenDest3+1:posDest4-1].strip()
                        Dest4=mapping[posDest4+lenDest4+1:posDest5-1].strip()
                        Dest5=mapping[posDest5+lenDest5+1:posCategoria-1].strip()
                        Categoria=mapping[posCategoria+lenCategoria+1:posSign-1].strip()
                        Formula1=None
                        Sign1=mapping[posSign+lenSign+1:posEBASign-1].strip()
                        EBASign1=mapping[posEBASign+lenEBASign+1:posCC-1].strip()
                        CC1=mapping[posCC+lenCC+1:lenMapping-1].strip()                    
                    else:
                        Conto=None
                        Dest2=None
                        Dest3=None
                        Dest4=None
                        Dest5=None
                        Categoria=None
                        Formula1=None
                        CC1=None
                        Sign1=None
                        EBASign1=None
                    if nome_sheet != 'Test_Formula' and nome_sheet != 'Macro' and EBASign1 is not None:
                        MappingCellInstance = MappingCell(
                        IdTabMapping=nome_sheet,
                        RigaMapping=cella.row,
                        ColonnaMapping=cella.column,
                        Cod_Conto=Conto,
                        Cod_Dest2=Dest2,
                        Cod_Dest3=Dest3,
                        Cod_Dest4=Dest4,
                        Cod_Dest5=Dest5,
                        Cod_Categoria=Categoria,
                        Formula=Formula1,
                        CC=CC1,
                        Sign=Sign1,
                        EBASign=EBASign1,
                        Coordinate=cella.coordinate
                        )
                        MappingList.append(MappingCellInstance)       
          

MappingTab = {
    'Id_Tab': [MappingCellInstance.IdTabMapping for MappingCellInstance in MappingList],
    'Row': [MappingCellInstance.RigaMapping for MappingCellInstance in MappingList],
    'Column': [MappingCellInstance.ColonnaMapping for MappingCellInstance in MappingList],
    'Cod_Conto': [MappingCellInstance.Cod_Conto for MappingCellInstance in MappingList],
    'Cod_Dest2': [MappingCellInstance.Cod_Dest2 for MappingCellInstance in MappingList],
    'Cod_Dest3': [MappingCellInstance.Cod_Dest3 for MappingCellInstance in MappingList],
    'Cod_Dest4': [MappingCellInstance.Cod_Dest4 for MappingCellInstance in MappingList],
    'Cod_Dest5': [MappingCellInstance.Cod_Dest5 for MappingCellInstance in MappingList],
    'Cod_Categoria': [MappingCellInstance.Cod_Categoria for MappingCellInstance in MappingList],
    'Formula': [MappingCellInstance.Formula for MappingCellInstance in MappingList],
    'Calculation_Logic': [MappingCellInstance.CC for MappingCellInstance in MappingList],
    'DB_Storage_Sign': [MappingCellInstance.Sign for MappingCellInstance in MappingList],
    'EBA_Sign': [MappingCellInstance.EBASign for MappingCellInstance in MappingList],
    'Coordinate': [MappingCellInstance.Coordinate for MappingCellInstance in MappingList]
}


df = pd.DataFrame(MappingTab)
# nome_file_csv5=r'C:\Users\federico.mangini\Downloads\provaMappingNum1.csv'
# df.to_csv(nome_file_csv5, index=False) 
MappingTab = df.to_dict(orient='records') 

# for record in MappingTab:
#     for record2 in DatiColonneTab:
#         if record.get('Id_Tab')==record2.get('IdTabCol') and record.get('Column')==record2.get('NumCol'):
#             record['Column']=record2.get('ValCol')
#     for record3 in DatiRigheTab:
#         if record.get('Id_Tab')==record3.get('IdTabRow') and record.get('Row')==record3.get('NumRow'):
#             record['Row']=record3.get('ValRow')
#             print(record)

Tab = []

for record in MappingTab:
    final=record.copy()
    for record2 in DatiColonneTab:
        if record['Id_Tab']==record2['IdTabCol'] and record['Column']==record2['NumCol']:
            tmp1=record2['ValCol']
            final['Column'] = "c" + str(tmp1).zfill(4)
    for record3 in DatiRigheTab:
        if record['Id_Tab']==record3['IdTabRow'] and record['Row']==record3['NumRow']:
            tmp1=record3['ValRow']
            final['Row'] = "r" + str(tmp1).zfill(4)     
    Tab.append(final)

df = pd.DataFrame(Tab)
nome_file_csv6= save_path + '\\' + start_name + filename + end_name + r'.csv'
df.to_csv(nome_file_csv6, index=False) 


# Registra l'ora di fine
end_time = datetime.datetime.now()
print("Ora di fine:", end_time)
duration = end_time - start_time
print("Durata dell'esecuzione:", duration)
                 
print('Codice concluso')

