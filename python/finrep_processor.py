"""Motore FinRep adattato per Pyodide.

La logica deriva dallo script desktop FinRepFlatDataModel.py, ma:
- riceve il workbook dal filesystem virtuale del browser;
- non usa percorsi Windows;
- non scrive direttamente sul disco dell'utente;
- restituisce record e CSV a JavaScript.
"""

import csv
import io
import json
import openpyxl

FIELDNAMES = [
    "Id_Tab", "Row", "Column", "Cod_Conto", "Cod_Dest2", "Cod_Dest3",
    "Cod_Dest4", "Cod_Dest5", "Cod_Categoria", "Formula",
    "Calculation_Logic", "DB_Storage_Sign", "EBA_Sign", "Coordinate",
]


def _rgb(cell):
    color = cell.fill.fgColor
    return getattr(color, "rgb", None)


def _extract_between(text, start_marker, end_marker):
    start = text.find(start_marker)
    end = text.find(end_marker, start + len(start_marker))
    if start == -1 or end == -1:
        return None
    # Mantiene l'offset usato dal motore originale.
    return text[start + len(start_marker):end].strip().strip("'").strip()


def _extract_last(text, marker):
    start = text.find(marker)
    if start == -1:
        return None
    return text[start + len(marker):].strip().strip("'").strip()


def _parse_mapping(raw_value):
    if raw_value is None:
        return None

    mapping = str(raw_value).replace("\n", "")
    account = "Account= "
    dest2 = "Dest 2 = "
    dest3 = "Dest 3 = "
    dest4 = "Dest 4 = "
    dest5 = "Dest 5 = "
    category = "Category= "
    formula = "Formula= "
    db_sign = "DB Storage Sign= "
    eba_sign = "EBA Sign= "
    calc_logic = "Calculation logic= "

    result = {
        "Cod_Conto": None, "Cod_Dest2": None, "Cod_Dest3": None,
        "Cod_Dest4": None, "Cod_Dest5": None, "Cod_Categoria": None,
        "Formula": None, "Calculation_Logic": None,
        "DB_Storage_Sign": None, "EBA_Sign": None,
    }

    has_account = account in mapping
    has_formula = formula in mapping
    has_dest2 = dest2 in mapping
    has_calc = calc_logic in mapping

    if has_account and not has_formula and has_dest2 and not has_calc:
        result.update({
            "Cod_Conto": _extract_between(mapping, account, dest2),
            "Cod_Dest2": _extract_between(mapping, dest2, dest3),
            "Cod_Dest3": _extract_between(mapping, dest3, dest4),
            "Cod_Dest4": _extract_between(mapping, dest4, dest5),
            "Cod_Dest5": _extract_between(mapping, dest5, category),
            "Cod_Categoria": _extract_between(mapping, category, db_sign),
            "DB_Storage_Sign": _extract_between(mapping, db_sign, eba_sign),
            "EBA_Sign": _extract_last(mapping, eba_sign),
        })
    elif not has_account and has_formula and not has_calc:
        result.update({
            "Formula": _extract_between(mapping, formula, db_sign),
            "DB_Storage_Sign": _extract_between(mapping, db_sign, eba_sign),
            "EBA_Sign": _extract_last(mapping, eba_sign),
        })
    elif has_account and not has_formula and has_dest2 and has_calc:
        result.update({
            "Cod_Conto": _extract_between(mapping, account, dest2),
            "Cod_Dest2": _extract_between(mapping, dest2, dest3),
            "Cod_Dest3": _extract_between(mapping, dest3, dest4),
            "Cod_Dest4": _extract_between(mapping, dest4, dest5),
            "Cod_Dest5": _extract_between(mapping, dest5, category),
            "Cod_Categoria": _extract_between(mapping, category, db_sign),
            "DB_Storage_Sign": _extract_between(mapping, db_sign, eba_sign),
            "EBA_Sign": _extract_between(mapping, eba_sign, calc_logic),
            "Calculation_Logic": _extract_last(mapping, calc_logic),
        })
    else:
        return None

    return result if result["EBA_Sign"] is not None else None


def process_finrep(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    sheet_names = workbook.sheetnames
    if not sheet_names:
        raise ValueError("Il file Excel non contiene fogli validi.")

    reference_sheet = workbook[sheet_names[0]]
    reference_style = reference_sheet["A1"]._style
    row_style = reference_sheet["A2"]._style
    column_style = reference_sheet["B1"]._style
    reference_color = _rgb(reference_sheet["A1"])

    columns = []
    rows = []
    max_columns = {}
    max_rows = {}

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        count_c = 1
        count_r = 1

        # Replica la prima scansione per righe dello script originale.
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                if cell.column == 1:
                    continue
                if cell.value is not None and (cell._style == column_style or _rgb(cell) == reference_color):
                    count_c += 1
                    columns.append({
                        "IdTabCol": sheet_name,
                        "ValCol": str(cell.value).zfill(4),
                        "NumCol": count_c,
                    })
                elif cell.value is None:
                    break
            break
        max_columns[sheet_name] = count_c

        # Replica la prima scansione per colonne dello script originale.
        for column_cells in sheet.iter_cols():
            for cell in column_cells:
                if cell.row == 1:
                    continue
                if cell.value is not None and (cell._style == row_style or _rgb(cell) == reference_color):
                    count_r += 1
                    rows.append({
                        "IdTabRow": sheet_name,
                        "ValRow": str(cell.value).zfill(4),
                        "NumRow": count_r,
                    })
                elif cell.value is None:
                    break
            break
        max_rows[sheet_name] = count_r

    column_lookup = {
        (item["IdTabCol"], item["NumCol"]): "c" + str(item["ValCol"]).zfill(4)
        for item in columns
    }
    row_lookup = {
        (item["IdTabRow"], item["NumRow"]): "r" + str(item["ValRow"]).zfill(4)
        for item in rows
    }

    records = []
    for sheet_name in sheet_names:
        if sheet_name in {"Test_Formula", "Macro"}:
            continue

        sheet = workbook[sheet_name]
        max_c = max_columns[sheet_name]
        max_r = max_rows[sheet_name]

        for row_cells in sheet.iter_rows(min_row=2, max_row=max_r, min_col=2, max_col=max_c):
            for cell in row_cells:
                if _rgb(cell) == reference_color:
                    continue
                parsed = _parse_mapping(cell.value)
                if parsed is None:
                    continue

                record = {
                    "Id_Tab": sheet_name,
                    "Row": row_lookup.get((sheet_name, cell.row), cell.row),
                    "Column": column_lookup.get((sheet_name, cell.column), cell.column),
                    **parsed,
                    "Coordinate": cell.coordinate,
                }
                records.append(record)

    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=FIELDNAMES, lineterminator="\n")
    writer.writeheader()
    writer.writerows(records)

    return json.dumps({
        "records": records,
        "csv": output.getvalue(),
        "count": len(records),
    }, ensure_ascii=False)
